"""
Export service — generates PDF and Excel files from composite rankings.

PDF:  WeasyPrint (CSS-based HTML → PDF)
Excel: openpyxl
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

_HEADERS = [
    "Rank",
    "Player",
    "Position",
    "Team",
    "PuckLogic Score",
    "Value Over Replacement",
    "OffNightGames",
    "Source Count",
    "G",
    "A",
    "PPP",
    "SOG",
    "Hits",
    "Blocks",
    "GP",
    "W",
    "GA",
    "SV%",
]

# Skater positions in display order, then goalies.
_POSITION_ORDER = ["C", "LW", "RW", "D", "G"]

_NOTES = [
    (
        "PuckLogic Score: The scoring-configuration-weighted fantasy point"
        " projection for this player."
    ),
    (
        "Value Over Replacement (VORP): Measures a player's projected fantasy"
        " value relative to the replacement-level player at their position in"
        " your league. Configure a league profile for this kit in your"
        " PuckLogic dashboard to unlock this column."
    ),
    "Source Count: The number of ranking sources that include this player.",
]


def _write_rankings_sheet(
    ws: Any,
    rankings: list[dict[str, Any]],
    header_fill: Any,
    header_font: Any,
) -> None:
    """Write headers and data rows onto *ws*."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rankings:
        stats = row.get("projected_stats", {})
        vorp = row.get("vorp")
        ws.append(
            [
                row["composite_rank"],
                row.get("name", ""),
                row.get("default_position", ""),
                row.get("team", ""),
                _fmt(row.get("projected_fantasy_points")),
                vorp if vorp is not None else "—",
                row.get("off_night_games", ""),
                row.get("source_count", 0),
                _fmt(stats.get("g")),
                _fmt(stats.get("a")),
                _fmt(stats.get("ppp")),
                _fmt(stats.get("sog")),
                _fmt(stats.get("hits")),
                _fmt(stats.get("blocks")),
                _fmt(stats.get("gp")),
                _fmt(stats.get("w")),
                _fmt(stats.get("ga")),
                _fmt(stats.get("sv_pct")),
            ]
        )

    for col_idx in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


def _write_by_position_sheet(
    ws: Any,
    rankings: list[dict[str, Any]],
    header_fill: Any,
    header_font: Any,
    section_font: Any,
) -> None:
    """Write a position-grouped sheet onto *ws*."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    by_position: dict[str, list[dict[str, Any]]] = {}
    for row in rankings:
        pos = row.get("default_position") or "?"
        by_position.setdefault(pos, []).append(row)

    ordered_positions = [p for p in _POSITION_ORDER if p in by_position]
    other_positions = [p for p in by_position if p not in _POSITION_ORDER]

    for pos in ordered_positions + other_positions:
        # Section header row
        ws.append([pos])
        section_row_idx = ws.max_row
        ws.cell(section_row_idx, 1).font = section_font

        for row in by_position[pos]:
            stats = row.get("projected_stats", {})
            vorp = row.get("vorp")
            ws.append(
                [
                    row["composite_rank"],
                    row.get("name", ""),
                    row.get("default_position", ""),
                    row.get("team", ""),
                    _fmt(row.get("projected_fantasy_points")),
                    vorp if vorp is not None else "—",
                    row.get("off_night_games", ""),
                    row.get("source_count", 0),
                    _fmt(stats.get("g")),
                    _fmt(stats.get("a")),
                    _fmt(stats.get("ppp")),
                    _fmt(stats.get("sog")),
                    _fmt(stats.get("hits")),
                    _fmt(stats.get("blocks")),
                    _fmt(stats.get("gp")),
                    _fmt(stats.get("w")),
                    _fmt(stats.get("ga")),
                    _fmt(stats.get("sv_pct")),
                ]
            )

    for col_idx in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


def generate_excel(
    rankings: list[dict[str, Any]],
    season: str,
    context_label: str | None = None,
) -> bytes:
    """Return an Excel workbook as bytes with two sheets.

    Sheet 1 — Full Rankings: all players sorted by fantasy points descending.
    Sheet 2 — By Position: same players grouped by position (C, LW, RW, D, G).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.properties.title = f"PuckLogic Rankings {season}"
    rendered_context_label = context_label or "scoring configuration"
    wb.properties.subject = (
        f"Source context: Source Count column; League context: {rendered_context_label}"
    )

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    section_font = Font(bold=True)

    ws1 = wb.active
    ws1.title = f"Full Rankings {season}"
    _write_rankings_sheet(ws1, rankings, header_fill, header_font)

    ws2 = wb.create_sheet(title="By Position")
    _write_by_position_sheet(ws2, rankings, header_fill, header_font, section_font)

    ws3 = wb.create_sheet(title="Notes")
    for note in _NOTES:
        ws3.append([note])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt(val: float | int | None) -> str | float | int:
    """Format a nullable numeric for the spreadsheet."""
    if val is None:
        return ""
    if isinstance(val, float):
        return round(val, 2)
    return val


_HTML_TEMPLATE = """\
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  h1 {{ color: #1e3a5f; font-size: 18px; }}
  .context {{ color: #475569; margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #1e3a5f; color: white; padding: 6px 8px; text-align: left; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #ddd; }}
  tr:nth-child(even) {{ background: #f4f7fb; }}
  .num {{ text-align: right; font-family: monospace; }}
  .rank {{ text-align: center; font-weight: bold; }}
  .footnote {{ color: #475569; font-style: italic; margin-top: 8px; font-size: 10px; }}
</style>
</head>
<body>
<h1>PuckLogic Draft Sheet</h1>
<p class="context">Season: {season}</p>
<p class="context">League context: {context_label}</p>
<p class="context">Generated: {generated_at}</p>
<table>
  <thead>
    <tr>
      <th>Rank</th><th>Player</th><th>Team</th><th>Pos</th>
      <th>PuckLogic Score</th><th>{vorp_header}</th><th>Off-Night</th>
      <th>G</th><th>A</th><th>PPP</th><th>SOG</th>
    </tr>
  </thead>
  <tbody>
    {rows}
  </tbody>
</table>
{footnote}
</body>
</html>
"""


def generate_pdf(
    rankings: list[dict[str, Any]],
    season: str,
    context_label: str | None = None,
    *,
    generated_at: str | None = None,
) -> bytes:
    """Return a PDF as bytes. Requires WeasyPrint to be installed."""
    from weasyprint import HTML

    rendered_generated_at = generated_at or datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    has_null_vorp = any(r.get("vorp") is None for r in rankings)
    vorp_header = "Value Over Replacement*" if has_null_vorp else "Value Over Replacement"
    footnote = (
        "<p class='footnote'>* Configure a league profile for this kit in your PuckLogic "
        "dashboard to unlock Value Over Replacement rankings.</p>"
        if has_null_vorp
        else ""
    )

    rows_html = ""
    for row in rankings:
        stats = row.get("projected_stats", {})
        fp = row.get("projected_fantasy_points")
        vorp = row.get("vorp")
        rows_html += (
            f"<tr>"
            f"<td class='rank'>{row['composite_rank']}</td>"
            f"<td>{row.get('name', '')}</td>"
            f"<td>{row.get('team', '')}</td>"
            f"<td>{row.get('default_position', '')}</td>"
            f"<td class='num'>{round(fp, 1) if fp is not None else '—'}</td>"
            f"<td class='num'>{round(vorp, 1) if vorp is not None else '—'}</td>"
            f"<td class='num'>{row.get('off_night_games', '—')}</td>"
            f"<td class='num'>{_fmt(stats.get('g'))}</td>"
            f"<td class='num'>{_fmt(stats.get('a'))}</td>"
            f"<td class='num'>{_fmt(stats.get('ppp'))}</td>"
            f"<td class='num'>{_fmt(stats.get('sog'))}</td>"
            f"</tr>\n"
        )

    html_content = _HTML_TEMPLATE.format(
        season=season,
        context_label=context_label or "scoring configuration",
        generated_at=rendered_generated_at,
        rows=rows_html,
        vorp_header=vorp_header,
        footnote=footnote,
    )
    return HTML(string=html_content).write_pdf()
