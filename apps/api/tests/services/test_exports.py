"""Unit tests for services/exports.py — generate_excel and generate_pdf.

generate_excel: tested for real (openpyxl is pure Python, no system deps).
generate_pdf:   WeasyPrint requires system libs (Cairo/Pango), so the entire
                weasyprint module is injected into sys.modules as a MagicMock;
                all HTML-construction logic is validated against the HTML string
                passed to the mock.
"""

from __future__ import annotations

import io
import sys
from typing import Any
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from services.exports import generate_excel, generate_pdf

# ---------------------------------------------------------------------------
# Shared fixtures / helpers
# ---------------------------------------------------------------------------

SEASON = "2025-26"

_EMPTY_STATS: dict[str, Any] = {
    "g": None,
    "a": None,
    "plus_minus": None,
    "pim": None,
    "ppg": None,
    "ppa": None,
    "ppp": None,
    "shg": None,
    "sha": None,
    "shp": None,
    "sog": None,
    "fow": None,
    "fol": None,
    "hits": None,
    "blocks": None,
    "gp": None,
    "gs": None,
    "w": None,
    "l": None,
    "ga": None,
    "sa": None,
    "sv": None,
    "sv_pct": None,
    "so": None,
    "otl": None,
}

PLAYER_A: dict[str, Any] = {
    "composite_rank": 1,
    "player_id": "p1",
    "name": "Connor McDavid",
    "team": "EDM",
    "default_position": "C",
    "platform_positions": ["C", "F"],
    "projected_fantasy_points": 290.0,
    "vorp": 50.0,
    "schedule_score": 0.8,
    "off_night_games": 10,
    "source_count": 3,
    "projected_stats": {**_EMPTY_STATS, "g": 60, "a": 90, "ppp": 50, "sog": 250, "gp": 82},
    "breakout_score": None,
    "regression_risk": None,
}

PLAYER_B: dict[str, Any] = {
    "composite_rank": 2,
    "player_id": "p2",
    "name": "Nathan MacKinnon",
    "team": "COL",
    "default_position": "C",
    "platform_positions": ["C"],
    "projected_fantasy_points": 265.0,
    "vorp": 30.0,
    "schedule_score": 0.7,
    "off_night_games": 8,
    "source_count": 2,
    "projected_stats": {**_EMPTY_STATS, "g": 50, "a": 80, "ppp": 40, "sog": 220, "gp": 80},
    "breakout_score": None,
    "regression_risk": None,
}

RANKINGS = [PLAYER_A, PLAYER_B]


def _make_weasyprint_mock(write_pdf_return: bytes = b"%PDF-1.4") -> MagicMock:
    """Return a fake weasyprint module whose HTML class behaves correctly."""
    mock_html_cls = MagicMock()
    mock_html_cls.return_value.write_pdf.return_value = write_pdf_return
    mock_module = MagicMock()
    mock_module.HTML = mock_html_cls
    return mock_module


def _capture_html(
    rankings: list[dict[str, Any]],
    season: str,
    *,
    generated_at: str | None = None,
) -> str:
    """Run generate_pdf with a mocked weasyprint and return the HTML string."""
    captured: dict[str, str] = {}

    def fake_html(*args: object, **kwargs: object) -> MagicMock:
        captured["html"] = str(kwargs.get("string", args[0] if args else ""))
        mock = MagicMock()
        mock.write_pdf.return_value = b"%PDF-1.4"
        return mock

    mock_module = MagicMock()
    mock_module.HTML.side_effect = fake_html

    with patch.dict(sys.modules, {"weasyprint": mock_module}):
        kwargs = {"generated_at": generated_at} if generated_at is not None else {}
        generate_pdf(rankings, season, **kwargs)

    return captured["html"]


# ---------------------------------------------------------------------------
# generate_excel — real openpyxl tests (no mocking needed)
# ---------------------------------------------------------------------------


class TestGenerateExcel:
    def test_context_label_is_in_workbook_subject(self) -> None:
        context_label = "Standard (sc-1); league profile: H2H; sources: hashtag:1"
        result = generate_excel(RANKINGS, SEASON, context_label)
        wb = load_workbook(io.BytesIO(result))
        assert context_label in (wb.properties.subject or "")

    def test_returns_bytes(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        assert isinstance(result, bytes)

    def test_bytes_are_valid_xlsx(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_worksheet_title_contains_season(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        assert SEASON in wb.active.title

    def test_worksheet_title_starts_with_full_rankings(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        assert wb.active.title.startswith("Full Rankings")

    def test_header_row_base_columns(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, 5)]
        assert headers == ["Rank", "Player", "Position", "Team"]

    def test_launch_xlsx_minimum_columns_and_context_metadata(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

        assert headers[:6] == [
            "Rank",
            "Player",
            "Position",
            "Team",
            "PuckLogic Score",
            "Projected Fantasy Value",
        ]
        assert "Source Count" in headers
        assert wb.properties.title == f"PuckLogic Rankings {SEASON}"
        assert "Source context" in (wb.properties.subject or "")

    def test_header_row_has_pucklogic_score(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        assert "PuckLogic Score" in headers

    def test_header_row_has_projected_fantasy_value(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        assert "Projected Fantasy Value" in headers

    def test_data_rows_count(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_first_data_row_rank_and_name(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "Connor McDavid"
        assert ws.cell(2, 3).value == "C"
        assert ws.cell(2, 4).value == "EDM"

    def test_first_data_row_fanpts(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        fp_col = headers.index("PuckLogic Score") + 1
        assert ws.cell(2, fp_col).value == pytest.approx(290.0, abs=0.01)

    def test_second_data_row_values(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.cell(3, 1).value == 2
        assert ws.cell(3, 2).value == "Nathan MacKinnon"
        assert ws.cell(3, 4).value == "COL"

    def test_single_player_no_sources(self) -> None:
        rankings = [
            {
                **PLAYER_A,
                "source_count": 0,
                "projected_fantasy_points": None,
            }
        ]
        result = generate_excel(rankings, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.max_row == 2  # header + 1 data row

    def test_empty_rankings_produces_header_only(self) -> None:
        result = generate_excel([], SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.max_row == 1
        assert ws.cell(1, 1).value == "Rank"

    def test_null_fp_writes_empty_string(self) -> None:
        rankings = [{**PLAYER_A, "projected_fantasy_points": None}]
        result = generate_excel(rankings, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        fp_col = headers.index("PuckLogic Score") + 1
        assert ws.cell(2, fp_col).value in (None, "")

    def test_header_uses_ga_not_gaa(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        assert "GA" in headers
        assert "GAA" not in headers

    def test_workbook_has_two_sheets(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 2

    def test_sheet2_title_is_by_position(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        assert wb.sheetnames[1] == "By Position"

    def test_sheet2_has_headers(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws2 = wb["By Position"]
        assert ws2.cell(1, 1).value == "Rank"
        assert ws2.cell(1, 2).value == "Player"

    def test_sheet2_contains_position_section_headers(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws2 = wb["By Position"]
        # Both players are C — expect a "C" section header row
        all_values = [ws2.cell(r, 1).value for r in range(1, ws2.max_row + 1)]
        assert "C" in all_values

    def test_sheet2_player_count_matches_sheet1(self) -> None:
        result = generate_excel(RANKINGS, SEASON)
        wb = load_workbook(io.BytesIO(result))
        ws1 = wb.active
        ws2 = wb["By Position"]
        # Sheet 1: 1 header + N players. Sheet 2: 1 header + M section headers + N players.
        # Player count on sheet2 = rows - 1 header - section header rows
        player_rows_sheet1 = ws1.max_row - 1
        # count non-section rows after header on sheet2 (rows where col1 is numeric)
        player_rows_sheet2 = sum(
            1 for r in range(2, ws2.max_row + 1) if isinstance(ws2.cell(r, 1).value, int)
        )
        assert player_rows_sheet2 == player_rows_sheet1


# ---------------------------------------------------------------------------
# generate_pdf — mocked WeasyPrint via sys.modules; validates HTML construction
# ---------------------------------------------------------------------------


class TestGeneratePdf:
    def test_returns_bytes(self) -> None:
        mock_module = _make_weasyprint_mock()
        with patch.dict(sys.modules, {"weasyprint": mock_module}):
            result = generate_pdf(RANKINGS, SEASON)
        assert isinstance(result, bytes)

    def test_returns_pdf_bytes_from_write_pdf(self) -> None:
        mock_module = _make_weasyprint_mock(b"%PDF-1.4 custom")
        with patch.dict(sys.modules, {"weasyprint": mock_module}):
            result = generate_pdf(RANKINGS, SEASON)
        assert result == b"%PDF-1.4 custom"

    def test_html_contains_season(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert SEASON in html

    def test_printable_draft_sheet_header_context_and_timestamp(self) -> None:
        html = _capture_html(
            RANKINGS,
            SEASON,
            generated_at="2026-05-11 15:30 UTC",
        )

        assert "PuckLogic Draft Sheet" in html
        assert f"Season: {SEASON}" in html
        assert "League context: scoring configuration" in html
        assert "Generated: 2026-05-11 15:30 UTC" in html
        assert "PuckLogic Score" in html
        assert "Projected Fantasy Value" in html

    def test_printable_draft_sheet_uses_passed_context_label(self) -> None:
        context_label = "Standard (sc-1); league profile: H2H; sources: hashtag:1"
        html = _capture_html(
            RANKINGS,
            SEASON,
            generated_at="2026-05-11 15:30 UTC",
        )

        assert "League context: scoring configuration" in html

        mock_module = _make_weasyprint_mock()
        with patch.dict(sys.modules, {"weasyprint": mock_module}):
            generate_pdf(RANKINGS, SEASON, context_label, generated_at="2026-05-11 15:30 UTC")

        rendered_html = mock_module.HTML.call_args.kwargs["string"]
        assert f"League context: {context_label}" in rendered_html

    def test_html_contains_player_names(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert "Connor McDavid" in html
        assert "Nathan MacKinnon" in html

    def test_html_contains_teams(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert "EDM" in html
        assert "COL" in html

    def test_html_contains_positions(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert "C" in html

    def test_html_contains_composite_ranks(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert ">1<" in html
        assert ">2<" in html

    def test_html_contains_pucklogic_score(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert "PuckLogic Score" in html

    def test_html_contains_projected_fantasy_value(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert "Projected Fantasy Value" in html

    def test_html_missing_fanpts_shows_em_dash(self) -> None:
        rankings = [{**PLAYER_A, "projected_fantasy_points": None}]
        html = _capture_html(rankings, SEASON)
        assert "—" in html

    def test_empty_rankings_produces_valid_html_structure(self) -> None:
        html = _capture_html([], SEASON)
        assert "<table>" in html
        assert SEASON in html
        assert "<tbody>" in html

    def test_html_is_valid_document_structure(self) -> None:
        html = _capture_html(RANKINGS, SEASON)
        assert "<!DOCTYPE html>" in html
        assert "<html>" in html
        assert "</html>" in html
        assert "<table>" in html
        assert "</table>" in html

    def test_weasyprint_html_called_with_string_kwarg(self) -> None:
        mock_module = _make_weasyprint_mock()
        with patch.dict(sys.modules, {"weasyprint": mock_module}):
            generate_pdf(RANKINGS, SEASON)
        call_kwargs = mock_module.HTML.call_args.kwargs
        assert "string" in call_kwargs

    def test_write_pdf_called_once(self) -> None:
        mock_module = _make_weasyprint_mock()
        with patch.dict(sys.modules, {"weasyprint": mock_module}):
            generate_pdf(RANKINGS, SEASON)
        mock_module.HTML.return_value.write_pdf.assert_called_once()
